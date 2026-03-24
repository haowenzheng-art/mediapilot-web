
"""
MediaPilot Excel导出模块
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


class ExcelExporter:
    """Excel导出器"""

    def __init__(self):
        self.header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_competitors(self, accounts):
        """导出对标账号到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "对标账号"

        headers = ["账号ID", "昵称", "平台", "粉丝数", "获赞数", "作品数", "平均点赞", "主页链接"]

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border

        # 写入数据
        for row_num, account in enumerate(accounts, 2):
            ws.cell(row=row_num, column=1, value=account.get("account_id", ""))
            ws.cell(row=row_num, column=2, value=account.get("nickname", ""))
            ws.cell(row=row_num, column=3, value=account.get("platform", ""))
            ws.cell(row=row_num, column=4, value=account.get("followers", 0))
            ws.cell(row=row_num, column=5, value=account.get("total_likes", 0))
            ws.cell(row=row_num, column=6, value=account.get("video_count", 0))
            ws.cell(row=row_num, column=7, value=account.get("avg_likes", 0))
            ws.cell(row=row_num, column=8, value=account.get("profile_url", ""))

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_transcript(self, video_title, lines):
        """导出逐字稿到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "逐字稿"

        ws.cell(row=1, column=1, value=video_title)
        ws.merge_cells('A1:B1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = ["时间", "文字"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_num, line in enumerate(lines, 3):
            ws.cell(row=row_num, column=1, value=line.get("time", ""))
            ws.cell(row=row_num, column=2, value=line.get("text", ""))

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_script(self, topic, script):
        """导出生成的脚本到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "分镜头脚本"

        ws.cell(row=1, column=1, value=f"选题: {topic}")
        ws.merge_cells('A1:E1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = ["场景", "时长", "画面", "台词", "备注"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_num, shot in enumerate(script, 3):
            ws.cell(row=row_num, column=1, value=shot.get("scene", ""))
            ws.cell(row=row_num, column=2, value=shot.get("duration", ""))
            ws.cell(row=row_num, column=3, value=shot.get("visual", ""))
            ws.cell(row=row_num, column=4, value=shot.get("audio", ""))
            ws.cell(row=row_num, column=5, value=shot.get("notes", ""))

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 25

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

